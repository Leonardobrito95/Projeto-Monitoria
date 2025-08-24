import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import customtkinter as ctk
from tkcalendar import DateEntry
import pandas as pd
import sqlite3
import os
from datetime import datetime
import matplotlib
matplotlib.use('TkAgg')
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
from PIL import Image
from dotenv import load_dotenv

# Carrega as variáveis de ambiente do arquivo .env
load_dotenv()

# Importa o novo módulo de análise
import analyzer

# --- CONFIGURAÇÕES DA APLICAÇÃO ---
DB_FILE = 'monitoria.db'
EXCEL_FILE = 'Métricas de Atendimento.xlsx'
ASSETS_DIR = 'assets'
APP_LOGO_FILE = os.path.join(ASSETS_DIR, 'logo_canaa.png')
APP_ICON_FILE = os.path.join(ASSETS_DIR, 'icon_canaa.png')
ADMIN_PASSWORD = os.getenv("MONITORIA_ADMIN_PASSWORD", "admin123")
COLUNAS = [
    'Motivo do Atendimento', 'Monitoria Zero', 'Protocolo', 'Data M', 'Nome do Agente', 'Equipe', 
    'Script inicial/final', 'Sondagem', 'Conhecimento técnico', 'Vícios de linguagem', 'Tom de voz', 
    'Cordialidade', 'Controle de Objeção', 'Ofensa Verbal', 'Retorno ao cliente', 'Ação de retenção',
    'Confirmação de dados', 'Transferencia Indevida', 'Uso do Mute', 'Erro de procedimento',
    'Negociação e venda', 'Inf. Protocolo?', 'Agilidade', 'Prontidão', 'Tabulação',
    'Resolução do conflito', 'Personalização', 'Omissão de atendimento', 'Avaliação ATD.',
    'Erro Crítico?', 'Itens Aplicáveis', 'Pontuação', 'Observações'
]
YES_NO_FIELDS = [
    'Script inicial/final', 'Sondagem', 'Conhecimento técnico', 'Vícios de linguagem', 'Tom de voz', 
    'Cordialidade', 'Controle de Objeção', 'Ofensa Verbal', 'Retorno ao cliente', 'Ação de retenção',
    'Confirmação de dados', 'Transferencia Indevida', 'Uso do Mute', 'Erro de procedimento',
    'Negociação e venda', 'Agilidade', 'Prontidão', 'Tabulação', 'Resolução do conflito', 
    'Personalização', 'Omissão de atendimento'
]
CRITICAL_ERRORS = {
    'Omissão de atendimento': 'Não Conforme', 'Ofensa Verbal': 'Não Conforme',
    'Erro de procedimento': 'Não Conforme', 'Confirmação de dados': 'Não Conforme',
    'Inf. Protocolo?': 'Não Conforme'
}

# Penalizações para cada critério (quando "Não Conforme")
PENALIZACOES = {
    'Script inicial/final': 0.50,
    'Sondagem': 0.50,
    'Conhecimento técnico': 0.50,
    'Vícios de linguagem': 0.50,
    'Transferencia Indevida': 0.50,
    'Ofensa Verbal' : 1.00,
    'Controle de Objeção': 0.50,
    'Retorno ao cliente': 0.50,
    'Ação de retenção': 0.50,
    'Confirmação de dados': 1.00,
    'Tom de voz': 0.50,
    'Uso do Mute': 0.50,
    'Erro de procedimento': 1.00,
    'Negociação e venda': 0.50,
    'Inf. Protocolo?': 1.00,
    'Agilidade': 0.50,
    'Prontidão': 0.50,
    'Tabulação': 0.50,
    'Resolução do conflito': 0.50,
    'Cordialidade': 0.50,
    'Personalização': 0.50,
    'Omissão de atendimento': 0.00
}

# Agentes e suas equipes (usado como seed inicial para o DB)
AGENTES_EQUIPE = {
    'Sarah Couto': 'SAC',
    'Matheus Henrique': 'SAC',
    'Larissa Santos': 'SAC',
    'Manoel Junior': 'SAC',
    'Andressa Costa': 'SAC',
    'Priscilla Rodrigues': 'SAC', 
    'Matheus Ferreira': 'SAC',
    'Hyrum Castro': 'SAC',
    'Rafael Vieira': 'SAC',
    'Aline Dias': 'SAC',
    'Livia Reis': 'SAC',
    'Walison Rodrigues': 'SAC',
    'Caique Abreu': 'SAC',
    'Rafael Brito': 'N2',
    'Ubiratan Sobrinho': 'N2',
    'Karoliny Lira': 'Retenção',
    'Anna Santos': 'SAC',
    'Clara Vieira':'SAC',
    'Camila Brito':'SAC'
}

# Estado para modo de edição
edit_mode = False
edit_id = None
botao_salvar = None
edit_agente_mode = False
agente_em_edicao = None
canvas_bar = None
canvas_pie = None
app_logo_img = None
chat_window = None  # Janela flutuante de chat
depto_filter_frame = None # Frame para checkboxes de depto no dashboard
chat_fab = None  # Botão flutuante para abrir chat

# --- FUNÇÕES DO BANCO DE DADOS ---
check_dept_vars = {}

def init_db():
    """Inicializa o banco de dados SQLite."""
    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()

        cursor.execute(f'''
            CREATE TABLE IF NOT EXISTS monitoria (
                id INTEGER PRIMARY KEY AUTOINCREMENT
            )
        ''')

        cursor.execute(f"PRAGMA table_info(monitoria)")
        existing_columns = [info[1] for info in cursor.fetchall()]

        for col in COLUNAS:
            if col not in existing_columns:
                try:
                    cursor.execute(f'ALTER TABLE monitoria ADD COLUMN "{col}" TEXT')
                except sqlite3.OperationalError:
                    pass
        conn.commit()

        # Cria a tabela de agentes para persistir agentes/equipes
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS agentes (
                nome TEXT PRIMARY KEY,
                equipe TEXT NOT NULL
            )
        ''')

        cursor.execute('SELECT COUNT(*) FROM agentes')
        total_agentes = cursor.fetchone()[0]
        if total_agentes == 0:
            for nome, equipe in AGENTES_EQUIPE.items():
                cursor.execute('INSERT OR IGNORE INTO agentes (nome, equipe) VALUES (?, ?)', (nome, equipe))
        conn.commit()

def carregar_dados_iniciais():
    """Carrega agentes e equipes da tabela 'agentes'."""
    try:
        with sqlite3.connect(DB_FILE) as conn:
            df = pd.read_sql_query('SELECT nome, equipe FROM agentes', conn)
        if not df.empty:
            agentes = sorted(df['nome'].astype(str).tolist())
            equipes = sorted(df['equipe'].astype(str).unique().tolist())
            return equipes, agentes
    except Exception:
        # Fallback para o dicionário hardcoded em caso de erro no DB
        agentes = sorted(AGENTES_EQUIPE.keys())
        equipes = sorted(set(AGENTES_EQUIPE.values()))
        return equipes, agentes

def verificar_protocolo_duplicado(protocolo, exclude_id=None):
    """Verifica se o protocolo já existe no banco, exceto para o ID em edição."""
    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        if exclude_id:
            cursor.execute("SELECT COUNT(*) FROM monitoria WHERE Protocolo = ? AND id != ?", (protocolo, exclude_id))
        else:
            cursor.execute("SELECT COUNT(*) FROM monitoria WHERE Protocolo = ?", (protocolo,))
        count = cursor.fetchone()[0]
    return count > 0

def calcular_pontuacao(dados):
    """Calcula a pontuação e itens aplicáveis com base nos dados do formulário."""
    pontuacao = 10.0
    itens_aplicaveis = 0
    erro_critico = False

    for campo, valor_critico in CRITICAL_ERRORS.items():
        if dados.get(campo) == valor_critico:
            erro_critico = True
            pontuacao = 0.0
            break

    if not erro_critico:
        for campo in YES_NO_FIELDS:
            if dados.get(campo) == 'Não Conforme':
                pontuacao -= PENALIZACOES.get(campo, 0)
            if dados.get(campo) in ['Conforme', 'Não Conforme']:
                itens_aplicaveis += 1

    return max(0, pontuacao), itens_aplicaveis, 'Sim' if erro_critico else 'Não'

def salvar_monitoria():
    """Salva ou atualiza uma monitoria no banco de dados."""
    global edit_mode, edit_id
    dados = {}

    # Campos que usam o método .get() padrão
    campos_get = YES_NO_FIELDS + [
        'Nome do Agente', 'Equipe', 'Motivo do Atendimento', 'Protocolo',
        'Avaliação ATD.', 'Erro Crítico?', 'Inf. Protocolo?'
    ]

    for col in COLUNAS:
        if col in widgets and widgets[col]:
            if col in campos_get:
                dados[col] = widgets[col].get()
            elif col == 'Data M':
                dados[col] = widgets[col].get_date().strftime('%d/%m/%Y')
            elif col == 'Observações':
                dados[col] = widgets[col].get("1.0", tk.END).strip()
            elif col == 'Monitoria Zero':
                dados[col] = widgets[col].get() if widgets[col].get() != 'Nenhum' else ''

    required_fields = ['Protocolo', 'Nome do Agente', 'Equipe']
    if not all(dados.get(field) for field in required_fields):
        messagebox.showwarning("Campos Obrigatórios", "Preencha Protocolo, Nome do Agente e Equipe.")
        return

    if dados.get('Avaliação ATD.'):
        try:
            valor_avaliacao = float(dados['Avaliação ATD.'])
            if not (0.0 <= valor_avaliacao <= 10.0):
                messagebox.showwarning("Entrada Inválida", "O campo Avaliação ATD. deve estar entre 0 e 10.")
                return
        except (ValueError, TypeError):
            messagebox.showwarning("Entrada Inválida", "O campo Avaliação ATD. deve ser numérico.")
            return

    if verificar_protocolo_duplicado(dados['Protocolo'], exclude_id=edit_id if edit_mode else None):
        messagebox.showwarning("Protocolo Duplicado", "Este número de protocolo já está registrado.")
        return

    pontuacao, itens_aplicaveis, erro_critico = calcular_pontuacao(dados)
    dados['Pontuação'] = f"{pontuacao:.2f}"
    dados['Itens Aplicáveis'] = str(itens_aplicaveis)
    dados['Erro Crítico?'] = erro_critico

    for col in COLUNAS:
        if col not in dados:
            dados[col] = ''

    try:
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            if edit_mode:
                columns = ', '.join([f'"{col}" = ?' for col in COLUNAS])
                values = [dados.get(col, '') for col in COLUNAS] + [edit_id]
                cursor.execute(f'UPDATE monitoria SET {columns} WHERE id = ?', values)
            else:
                columns = ', '.join([f'"{col}"' for col in COLUNAS])
                placeholders = ', '.join(['?' for _ in COLUNAS])
                values = [dados.get(col, '') for col in COLUNAS]
                cursor.execute(f'INSERT INTO monitoria ({columns}) VALUES ({placeholders})', values)
            conn.commit()

        update_excel()
        messagebox.showinfo("Sucesso", "Monitoria salva com sucesso!" if not edit_mode else "Monitoria atualizada com sucesso!")
        limpar_formulario()
        aplicar_filtros()
        aplicar_filtros_dashboard()
        if edit_mode:
            edit_mode = False
            edit_id = None
            botao_salvar.configure(text="Salvar Monitoria")
            tabview.set("Nova Monitoria")
    except Exception as e:
        messagebox.showerror("Erro ao Salvar", f"Erro ao salvar dados: {e}")

def update_excel():
    """Atualiza a aba 'Base de dados da Monitoria' no arquivo Excel."""
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        with sqlite3.connect(DB_FILE) as conn:
            df = pd.read_sql_query("SELECT * FROM monitoria", conn)
        if 'id' in df.columns:
            df = df.drop(columns=['id'])

        df['Data M'] = pd.to_datetime(df['Data M'], format='%d/%m/%Y', errors='coerce').dt.date
        for col in ['Avaliação ATD.', 'Itens Aplicáveis', 'Pontuação']:
            df[col] = pd.to_numeric(df[col], errors='coerce').round(2).fillna(0)

        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
        else:
            from openpyxl import Workbook
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

        if 'Base de dados da Monitoria' in wb.sheetnames:
            wb.remove(wb['Base de dados da Monitoria'])

        ws = wb.create_sheet('Base de dados da Monitoria')

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.fill = PatternFill(start_color='4A90E2', end_color='4A90E2', fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for col_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = max(max_length + 2, 12)

        wb.save(EXCEL_FILE)

    except Exception as e:
        messagebox.showerror("Erro ao Atualizar Excel", f"Erro ao atualizar Excel: {e}")

def excluir_registro():
    """Exclui o registro selecionado após confirmação."""
    selected_items = tree.selection()
    if not selected_items:
        messagebox.showwarning("Nenhum Registro Selecionado", "Selecione um registro para excluir.")
        return
    
    selected_item = selected_items[0]
    values = tree.item(selected_item, 'values')
    protocolo_selecionado = values[COLUNAS.index('Protocolo')]
    try:
        registro_id = int(selected_item)
    except ValueError:
        registro_id = None

    if not messagebox.askyesno("Confirmar Exclusão", f"Deseja realmente excluir a monitoria com protocolo {protocolo_selecionado}?"):
        return

    try:
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            if registro_id is not None:
                cursor.execute("DELETE FROM monitoria WHERE id = ?", (registro_id,))
            else:
                # Fallback caso o iid não seja um id numérico
                cursor.execute("DELETE FROM monitoria WHERE Protocolo = ?", (protocolo_selecionado,))
            conn.commit()

        update_excel()
        aplicar_filtros()
        aplicar_filtros_dashboard()
        messagebox.showinfo("Sucesso", f"Monitoria com protocolo {protocolo_selecionado} excluída com sucesso!")
    except Exception as e:
        messagebox.showerror("Erro ao Excluir", f"Erro ao excluir registro: {e}")

def limpar_formulario():
    """Limpa todos os campos do formulário."""
    global edit_mode, edit_id
    for col, widget in widgets.items():
        if not widget:
            continue
        if col in YES_NO_FIELDS:
            widget.set('Conforme')
            if col in CRITICAL_ERRORS:
                widget.configure(text_color='#FFFFFF')
        elif col == 'Erro Crítico?':
            widget.set('Não')
            widget.configure(text_color='#FFFFFF')
        elif col == 'Data M':
            widget.set_date(datetime.now())
        elif col in ['Protocolo', 'Avaliação ATD.']:
            widget.delete(0, 'end')
        elif col == 'Observações':
            widget.delete("1.0", tk.END)
        elif col in ['Motivo do Atendimento', 'Nome do Agente', 'Equipe']:
            widget.set('')
        elif col == 'Monitoria Zero':
            widget.set('Nenhum')

    widgets['Protocolo'].focus()
    if edit_mode:
        edit_mode = False
        edit_id = None
        botao_salvar.configure(text="Salvar Monitoria")

def atualizar_ultimos_lancamentos(filtro_agente=None, filtro_protocolo=None):
    """Atualiza a tabela com registros do banco de dados, aplicando filtros."""
    for i in tree.get_children():
        tree.delete(i)
    try:
        with sqlite3.connect(DB_FILE) as conn:
            query = "SELECT * FROM monitoria"
            conditions, params = [], []

            if filtro_agente and filtro_agente != "Todos":
                conditions.append('"Nome do Agente" = ?')
                params.append(filtro_agente)
            if filtro_protocolo:
                conditions.append('Protocolo LIKE ?')
                params.append(f'%{filtro_protocolo}%')

            if conditions:
                query += " WHERE " + " AND ".join(conditions)
            
            query += " ORDER BY id DESC"
            df = pd.read_sql_query(query, conn, params=params)
            
            for col in COLUNAS:
                if col not in df.columns:
                    df[col] = ''
            df = df.fillna('')

        tree['columns'] = COLUNAS
        for col in COLUNAS:
            tree.heading(col, text=col)
            tree.column(col, width=120 if col != 'Observações' else 200, anchor='center', stretch=tk.NO)

        for _, row in df.iterrows():
            iid = None
            if 'id' in df.columns and pd.notna(row['id']):
                try:
                    iid = str(int(row['id']))
                except (ValueError, TypeError):
                    iid = None
            valores = [row.get(col, '') for col in COLUNAS]
            tree.insert("", "end", iid=iid, values=valores)
    except Exception as e:
        messagebox.showerror("Erro de Leitura", f"Erro ao carregar lançamentos: {e}")

def aplicar_filtros():
    """Aplica filtros de agente e protocolo à tabela."""
    agente = combo_filtro_agente.get()
    protocolo = entry_filtro_protocolo.get().strip()
    atualizar_ultimos_lancamentos(filtro_agente=agente, filtro_protocolo=protocolo)

def limpar_filtros():
    """Limpa os campos de filtro e recarrega todos os registros."""
    combo_filtro_agente.set("Todos")
    entry_filtro_protocolo.delete(0, tk.END)
    atualizar_ultimos_lancamentos()

def atualizar_equipe(*args):
    """Atualiza o campo Equipe com base no agente selecionado."""
    agente = widgets['Nome do Agente'].get()
    equipe = ''
    try:
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT equipe FROM agentes WHERE nome = ?', (agente,))
            row = cursor.fetchone()
            if row:
                equipe = row[0]
    except Exception:
        # Fallback para o dicionário se o DB falhar
        equipe = AGENTES_EQUIPE.get(agente, '')
    widgets['Equipe'].set(equipe)

def atualizar_cor_critica(widget, campo):
    """Atualiza a cor do texto do ComboBox com base na seleção crítica."""
    valor = widget.get()
    if campo in CRITICAL_ERRORS and valor == CRITICAL_ERRORS[campo]:
        widget.configure(text_color="#FF0000")
    else:
        widget.configure(text_color='#FFFFFF')

def verificar_senha_admin():
    """Solicita e verifica a senha administrativa."""
    dialog = ctk.CTkInputDialog(title="Autenticação Administrativa", text="Digite a senha administrativa:")
    senha = dialog.get_input()
    return senha == ADMIN_PASSWORD if senha is not None else False

# --- AUXILIARES DE DATA (filtros) ---
def _parse_date_str(date_str):
    """Converte dd/mm/YYYY em objeto date. Retorna None se vazio ou inválido.""" # noqa
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, '%d/%m/%Y').date()
    except Exception:
        return None

def _to_ymd(date_obj):
    """Retorna YYYYMMDD para comparações lexicográficas no SQLite.""" # noqa
    if not date_obj:
        return None
    return date_obj.strftime('%Y%m%d')

def _atualizar_comboboxes_agentes():
    """Atualiza todos os comboboxes de agentes e equipes na UI."""
    equipes, agentes = carregar_dados_iniciais()
    widgets['Nome do Agente'].configure(values=agentes)
    combo_filtro_agente.configure(values=["Todos"] + agentes)
    combo_filtro_agente_dashboard.configure(values=["Todos"] + agentes)
    combo_filtro_equipe_dashboard.configure(values=["Todas"] + equipes)
    combo_equipe_novo_agente.configure(values=equipes)

def adicionar_agente():
    """Adiciona ou edita um agente no banco de dados e atualiza a interface."""
    global edit_agente_mode, agente_em_edicao
    if not verificar_senha_admin():
        messagebox.showerror("Erro de Autenticação", "Senha administrativa incorreta.")
        return

    nome_agente = entry_novo_agente.get().strip()
    equipe = combo_equipe_novo_agente.get()

    if not nome_agente or not equipe:
        messagebox.showwarning("Campos Vazios", "Por favor, insira o nome e a equipe do agente.")
        return

    try:
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            # Verifica se o agente já existe, ignorando o agente atual em modo de edição
            if edit_agente_mode and nome_agente.lower() == agente_em_edicao.lower():
                pass # Permite salvar alteração de equipe para o mesmo agente
            else:
                cursor.execute('SELECT COUNT(*) FROM agentes WHERE lower(nome) = ?', (nome_agente.lower(),))
                if cursor.fetchone()[0] > 0:
                    messagebox.showwarning("Agente Duplicado", f"O agente {nome_agente} já existe.")
                    return

            message = ""
            if edit_agente_mode:
                # Se o nome mudou, atualiza a chave primária (requer delete e insert)
                if agente_em_edicao.lower() != nome_agente.lower():
                    cursor.execute('DELETE FROM agentes WHERE nome = ?', (agente_em_edicao,))
                cursor.execute('INSERT OR REPLACE INTO agentes (nome, equipe) VALUES (?, ?)', (nome_agente, equipe))
                message = f"Agente {nome_agente} atualizado com sucesso!"
            else:
                cursor.execute('INSERT INTO agentes (nome, equipe) VALUES (?, ?)', (nome_agente, equipe))
                message = f"Agente {nome_agente} adicionado com sucesso!"
            conn.commit()

        _atualizar_comboboxes_agentes()

        listbox_agentes.delete(0, tk.END)
        with sqlite3.connect(DB_FILE) as conn:
            df = pd.read_sql_query('SELECT nome, equipe FROM agentes ORDER BY nome COLLATE NOCASE', conn)
        for _, r in df.iterrows():
            listbox_agentes.insert(tk.END, f"{r['nome']} ({r['equipe']})")

        entry_novo_agente.delete(0, tk.END)
        combo_equipe_novo_agente.set('')
        if edit_agente_mode:
            edit_agente_mode = False
            agente_em_edicao = None
            botao_adicionar_agente.configure(text="Adicionar Agente")

        messagebox.showinfo("Sucesso", message)

    except Exception as e:
        messagebox.showerror("Erro no Banco de Dados", f"Não foi possível salvar o agente: {e}")

def excluir_agente():
    """Exclui o agente selecionado do banco de dados e atualiza a interface."""
    if not verificar_senha_admin():
        messagebox.showerror("Erro de Autenticação", "Senha administrativa incorreta.")
        return

    try:
        selected = listbox_agentes.get(listbox_agentes.curselection())
        agente = selected.split(' (')[0]
    except tk.TclError:
        messagebox.showwarning("Nenhum Agente Selecionado", "Selecione um agente para excluir.")
        return

    if not messagebox.askyesno("Confirmar Exclusão", f"Deseja realmente excluir o agente {agente}?"):
        return

    try:
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM monitoria WHERE \"Nome do Agente\" = ?", (agente,))
            count = cursor.fetchone()[0]
            if count > 0:
                messagebox.showwarning("Agente em Uso", f"O agente {agente} está vinculado a {count} monitoria(s) e não pode ser excluído.")
                return

            cursor.execute('DELETE FROM agentes WHERE nome = ?', (agente,))
            conn.commit()

            _atualizar_comboboxes_agentes()

            # Recarrega a lista
            listbox_agentes.delete(0, tk.END)
            df = pd.read_sql_query('SELECT nome, equipe FROM agentes ORDER BY nome COLLATE NOCASE', conn)
            for _, r in df.iterrows():
                listbox_agentes.insert(tk.END, f"{r['nome']} ({r['equipe']})")

            messagebox.showinfo("Sucesso", f"Agente {agente} excluído com sucesso!")

    except Exception as e:
        messagebox.showerror("Erro no Banco de Dados", f"Não foi possível excluir o agente: {e}")

def editar_agente():
    """Carrega o agente selecionado para edição."""
    global edit_agente_mode, agente_em_edicao
    if not verificar_senha_admin():
        messagebox.showerror("Erro de Autenticação", "Senha administrativa incorreta.")
        return

    try:
        selected = listbox_agentes.get(listbox_agentes.curselection())
        agente_em_edicao = selected.split(' (')[0]
    except tk.TclError:
        messagebox.showwarning("Nenhum Agente Selecionado", "Selecione um agente para editar.")
        return

    entry_novo_agente.delete(0, tk.END)
    entry_novo_agente.insert(0, agente_em_edicao)
    try:
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT equipe FROM agentes WHERE nome = ?', (agente_em_edicao,))
            row = cursor.fetchone()
            combo_equipe_novo_agente.set(row[0] if row else '')
    except Exception:
        messagebox.showerror("Erro no Banco de Dados", f"Não foi possível carregar dados do agente: {e}")
        # Fallback para o dicionário se o DB falhar
        combo_equipe_novo_agente.set(AGENTES_EQUIPE.get(agente_em_edicao, ''))

    edit_agente_mode = True
    botao_adicionar_agente.configure(text="Salvar Alterações")

def alterar_senha_admin():
    """Altera a senha administrativa após validação."""
    global ADMIN_PASSWORD
    if not verificar_senha_admin():
        messagebox.showerror("Erro", "Senha administrativa atual incorreta.")
        return

    dialog_nova = ctk.CTkInputDialog(title="Alterar Senha", text="Digite a nova senha administrativa:")
    nova_senha = dialog_nova.get_input()

    if nova_senha:
        ADMIN_PASSWORD = nova_senha
        messagebox.showinfo("Sucesso", "Senha administrativa alterada com sucesso!")
    else:
        messagebox.showwarning("Ação Cancelada", "A nova senha não pode estar vazia.")

def limpar_lancamentos():
    """Limpa todos os lançamentos do banco de dados após validação administrativa."""
    if not verificar_senha_admin():
        messagebox.showerror("Erro de Autenticação", "Senha administrativa incorreta.")
        return

    if not messagebox.askyesno("Confirmar Limpeza", "Deseja realmente excluir TODOS os lançamentos? Esta ação não pode ser desfeita."):
        return

    try:
        with sqlite3.connect(DB_FILE) as conn:
            conn.execute("DELETE FROM monitoria")
            conn.commit()

        update_excel()
        aplicar_filtros()
        aplicar_filtros_dashboard()
        messagebox.showinfo("Sucesso", "Todos os lançamentos foram excluídos.")
    except Exception as e:
        messagebox.showerror("Erro ao Limpar", f"Erro ao limpar lançamentos: {e}")

def _atualizar_checkboxes_departamentos():
    """Atualiza os checkboxes de departamento na aba Dashboard."""
    global check_dept_vars, depto_filter_frame

    if not depto_filter_frame:
        return

    # Limpa o frame antigo
    for widget in depto_filter_frame.winfo_children():
        widget.destroy()
    check_dept_vars.clear()

    ctk.CTkLabel(depto_filter_frame, text="Departamentos para Auditoria IA:", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=(0, 10))

    try:
        # Usa o mapeamento que já é cacheado
        departamentos = analyzer.get_department_mapping().keys()
        for depto in sorted(list(departamentos)):
            check_dept_vars[depto] = tk.IntVar(value=0)
            ctk.CTkCheckBox(
                depto_filter_frame,
                text=depto,
                variable=check_dept_vars[depto],
                onvalue=1,
                offvalue=0
            ).pack(side="left", padx=5)
    except analyzer.APIError as e:
        # Se falhar, mostra um aviso e não exibe checkboxes
        ctk.CTkLabel(depto_filter_frame, text="Não foi possível carregar os departamentos da API.", text_color="orange").pack(side="left")
        messagebox.showwarning("API Desconectada", f"Não foi possível carregar a lista de departamentos para a auditoria IA.\n\n{e}")
    except Exception as e:
        ctk.CTkLabel(depto_filter_frame, text="Erro ao carregar departamentos.", text_color="red").pack(side="left")
        messagebox.showerror("Erro", f"Erro inesperado ao carregar departamentos: {e}")

def popular_lista_departamentos():
    """Carrega e exibe a lista de departamentos na aba de Configurações."""
    listbox_deptos.delete(0, tk.END)
    try:
        departamentos = analyzer.list_departments()
        for depto in sorted(departamentos, key=lambda d: d.get('nome', '')):
            listbox_deptos.insert(tk.END, f"{depto.get('nome')} (ID: {depto.get('_id')})")
    except analyzer.APIError as e:
        listbox_deptos.insert(tk.END, "Erro ao carregar departamentos.")
        messagebox.showerror("Erro de API", f"Não foi possível listar os departamentos:\n\n{e}")

def adicionar_departamento():
    """Cria um novo departamento via API."""
    nome_depto = entry_novo_depto.get().strip()
    if not nome_depto:
        messagebox.showwarning("Campo Vazio", "Por favor, insira o nome do departamento.")
        return
    try:
        analyzer.create_department(nome_depto)
        messagebox.showinfo("Sucesso", f"Departamento '{nome_depto}' criado com sucesso!")
        entry_novo_depto.delete(0, tk.END)
        popular_lista_departamentos()
        _atualizar_checkboxes_departamentos()
    except analyzer.APIError as e:
        messagebox.showerror("Erro na API", f"Não foi possível criar o departamento:\n\n{e}")

def atualizar_dashboard(filtro_agente=None, filtro_equipe=None, filtro_avaliacao=None, filtro_pontuacao=None, data_ini=None, data_fim=None):
    """Atualiza a aba Dashboard com métricas."""
    global canvas_bar, canvas_pie
    
    for i in dashboard_tree.get_children():
        dashboard_tree.delete(i)
    
    default_dashboard_columns = ['Agente', 'Média Pontuação', 'Total Monitorias', 'Erros Críticos', 'Média Erro Crítico (%)', 'Média Conforme (%)', 'Média Não Conforme (%)']
    dashboard_tree['columns'] = default_dashboard_columns
    for col in default_dashboard_columns:
        dashboard_tree.heading(col, text=col)
        dashboard_tree.column(col, width=150, anchor='center', stretch=tk.NO)

    try:
        with sqlite3.connect(DB_FILE) as conn:
            query = "SELECT * FROM monitoria"
            conditions, params = [], []

            if filtro_agente and filtro_agente != "Todos":
                conditions.append('"Nome do Agente" = ?')
                params.append(filtro_agente)
            if filtro_equipe and filtro_equipe != "Todas":
                conditions.append('Equipe = ?')
                params.append(filtro_equipe)
            if filtro_avaliacao:
                conditions.append('"Avaliação ATD." = ?')
                params.append(str(filtro_avaliacao))
            if filtro_pontuacao:
                conditions.append('Pontuação = ?')
                params.append(str(filtro_pontuacao))

            # intervalo de datas (Data M é armazenada como dd/mm/YYYY)
            # Convertendo para YYYYMMDD para comparação no SQLite
            expr_ymd = "substr(\"Data M\",7,4) || substr(\"Data M\",4,2) || substr(\"Data M\",1,2)"
            if data_ini:
                conditions.append(f"{expr_ymd} >= ?")
                params.append(_to_ymd(data_ini))
            if data_fim:
                conditions.append(f"{expr_ymd} <= ?")
                params.append(_to_ymd(data_fim))

            if conditions:
                query += " WHERE " + " AND ".join(conditions)
            
            df = pd.read_sql_query(query, conn, params=params)
        
        if df.empty:
            if canvas_bar: canvas_bar.get_tk_widget().destroy()
            if canvas_pie: canvas_pie.get_tk_widget().destroy()
            canvas_bar, canvas_pie = None, None
            return

        df['Pontuação'] = pd.to_numeric(df['Pontuação'], errors='coerce')
        df['Itens Aplicáveis'] = pd.to_numeric(df['Itens Aplicáveis'], errors='coerce')
        
        df['Total Conforme'] = df[YES_NO_FIELDS].eq('Conforme').sum(axis=1)
        df['Total Não Conforme'] = df[YES_NO_FIELDS].eq('Não Conforme').sum(axis=1)
        df['Total Itens Validos'] = df[YES_NO_FIELDS].isin(['Conforme', 'Não Conforme']).sum(axis=1)
        
        grouped = df.groupby('Nome do Agente')
        metrics = grouped.agg({
            'Pontuação': 'mean',
            'Protocolo': 'count',
            'Erro Crítico?': lambda x: (x == 'Sim').sum(),
            'Total Conforme': 'sum',
            'Total Não Conforme': 'sum',
            'Total Itens Validos': 'sum'
        }).reset_index()
        metrics.columns = ['Agente', 'Média Pontuação', 'Total Monitorias', 'Erros Críticos', 'Total Conforme', 'Total Não Conforme', 'Total Itens Validos']

        # CORREÇÃO: Tratamento de divisão por zero
        metrics['Média Conforme (%)'] = (metrics['Total Conforme'] / metrics['Total Itens Validos'] * 100).fillna(0).round(2)
        metrics['Média Não Conforme (%)'] = (metrics['Total Não Conforme'] / metrics['Total Itens Validos'] * 100).fillna(0).round(2)
        metrics['Média Erro Crítico (%)'] = (metrics['Erros Críticos'] / metrics['Total Monitorias'] * 100).fillna(0).round(2)
        
        metrics['Média Pontuação'] = pd.to_numeric(metrics['Média Pontuação'], errors='coerce').fillna(0).round(2)
        
        display_metrics = metrics[['Agente', 'Média Pontuação', 'Total Monitorias', 'Erros Críticos', 'Média Erro Crítico (%)', 'Média Conforme (%)', 'Média Não Conforme (%)']]
        
        for _, row in display_metrics.iterrows():
            row_values = list(row)
            row_values[1] = f"{row['Média Pontuação']:.2f}"
            dashboard_tree.insert("", "end", values=row_values)
        
        update_charts(metrics, filtro_agente, filtro_equipe, data_ini, data_fim)

    except Exception as e:
        messagebox.showerror("Erro ao Carregar Dashboard", f"Erro ao carregar dashboard: {e}")

def update_charts(df, filtro_agente, filtro_equipe, data_ini=None, data_fim=None):
    """Atualiza os gráficos de barras e pizza na aba Dashboard."""
    global canvas_bar, canvas_pie
    
    if canvas_bar: canvas_bar.get_tk_widget().destroy()
    if canvas_pie: canvas_pie.get_tk_widget().destroy()
    canvas_bar, canvas_pie = None, None
    
    if df.empty: return
    
    # Gráfico de barras
    fig_bar, ax_bar = plt.subplots(figsize=(6, 4))
    ax_bar.bar(df['Agente'], pd.to_numeric(df['Média Pontuação'], errors='coerce').fillna(0), color='#4A90E2')
    ax_bar.set_title('Média de Pontuação por Agente', color='white')
    ax_bar.set_xlabel('Agente', color='white')
    ax_bar.set_ylabel('Pontuação', color='white')
    ax_bar.set_ylim(0, 10)
    ax_bar.tick_params(axis='x', rotation=45, colors='white')
    ax_bar.tick_params(axis='y', colors='white')
    fig_bar.patch.set_facecolor('#2a2d2e')
    ax_bar.set_facecolor('#2a2d2e')
    plt.tight_layout()
    
    canvas_bar = FigureCanvasTkAgg(fig_bar, master=charts_frame)
    canvas_bar.draw()
    canvas_bar.get_tk_widget().pack(side='left', padx=5, pady=5, fill='both', expand=True)
    plt.close(fig_bar)
    
    # Gráfico de pizza
    with sqlite3.connect(DB_FILE) as conn:
        query = 'SELECT "Erro Crítico?", COUNT(*) as count FROM monitoria'
        conditions, params = [], []
        if filtro_agente and filtro_agente != "Todos":
            conditions.append('"Nome do Agente" = ?')
            params.append(filtro_agente)
        if filtro_equipe and filtro_equipe != "Todas":
            conditions.append('Equipe = ?')
            params.append(filtro_equipe)
        expr_ymd = "substr(\"Data M\",7,4) || substr(\"Data M\",4,2) || substr(\"Data M\",1,2)"
        if data_ini:
            conditions.append(f"{expr_ymd} >= ?")
            params.append(_to_ymd(data_ini))
        if data_fim:
            conditions.append(f"{expr_ymd} <= ?")
            params.append(_to_ymd(data_fim))
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
        query += ' GROUP BY "Erro Crítico?"'
        df_pie = pd.read_sql_query(query, conn, params=params)
    
    if not df_pie.empty:
        labels = df_pie['Erro Crítico?']
        sizes = df_pie['count']
        colors = ['#FF4C4C' if label == 'Sim' else '#4A90E2' for label in labels]
        explode = [0.1 if label == 'Sim' else 0 for label in labels]
        
        fig_pie, ax_pie = plt.subplots(figsize=(4, 4))
        ax_pie.pie(sizes, explode=explode, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90, textprops={'color':"w"})
        ax_pie.set_title('Proporção de Erros Críticos', color='white')
        fig_pie.patch.set_facecolor('#2a2d2e')
        plt.tight_layout()
        
        canvas_pie = FigureCanvasTkAgg(fig_pie, master=charts_frame)
        canvas_pie.draw()
        canvas_pie.get_tk_widget().pack(side='left', padx=5, pady=5, fill='both', expand=True)
        plt.close(fig_pie)

def aplicar_filtros_dashboard():
    """Aplica filtros ao Dashboard."""
    agente = combo_filtro_agente_dashboard.get()
    equipe = combo_filtro_equipe_dashboard.get()
    avaliacao = entry_filtro_avaliacao_dashboard.get().strip()
    pontuacao = entry_filtro_pontuacao_dashboard.get().strip()
    data_ini = entry_data_ini_dashboard.get_date() if entry_data_ini_dashboard.get() else None
    data_fim = entry_data_fim_dashboard.get_date() if entry_data_fim_dashboard.get() else None
    if data_ini and data_fim and data_ini > data_fim:
        messagebox.showwarning("Período inválido", "A data inicial não pode ser maior que a data final.")
        return
    atualizar_dashboard(
        filtro_agente=agente if agente != "Todos" else None,
        filtro_equipe=equipe if equipe != "Todas" else None,
        filtro_avaliacao=avaliacao,
        filtro_pontuacao=pontuacao,
        data_ini=data_ini,
        data_fim=data_fim
    )

def limpar_filtros_dashboard():
    """Limpa os filtros do Dashboard."""
    combo_filtro_agente_dashboard.set("Todos")
    combo_filtro_equipe_dashboard.set("Todas")
    entry_filtro_avaliacao_dashboard.delete(0, tk.END)
    entry_filtro_pontuacao_dashboard.delete(0, tk.END)
    try:
        entry_data_ini_dashboard.set_date(datetime.now().replace(day=1))
        entry_data_fim_dashboard.set_date(datetime.now())
    except Exception:
        pass
    atualizar_dashboard()

def gerar_relatorio():
    """Gera um relatório em Excel com os dados atuais."""
    default_dashboard_columns = ['Agente', 'Média Pontuação', 'Total Monitorias', 'Erros Críticos', 'Média Erro Crítico (%)', 'Média Conforme (%)', 'Média Não Conforme (%)']
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f"Relatorio_Monitoria_{timestamp}.xlsx"
        output_excel = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=default_filename,
            title="Salvar Relatório Como"
        )
        if not output_excel: return

        # Obter dados do dashboard da Treeview (reflete filtros)
        dashboard_data = [dashboard_tree.item(item, 'values') for item in dashboard_tree.get_children()]
        df_dashboard = pd.DataFrame(dashboard_data, columns=default_dashboard_columns)
        df_dashboard['Média Pontuação'] = pd.to_numeric(df_dashboard['Média Pontuação'], errors='coerce').fillna(0)

        # Obter todos os lançamentos do banco de dados respeitando os filtros atuais do dashboard
        agente = combo_filtro_agente_dashboard.get()
        equipe = combo_filtro_equipe_dashboard.get()
        avaliacao = entry_filtro_avaliacao_dashboard.get().strip()
        pontuacao = entry_filtro_pontuacao_dashboard.get().strip()
        data_ini = entry_data_ini_dashboard.get_date() if entry_data_ini_dashboard.get() else None
        data_fim = entry_data_fim_dashboard.get_date() if entry_data_fim_dashboard.get() else None

        with sqlite3.connect(DB_FILE) as conn:
            query = "SELECT * FROM monitoria"
            conditions, params = [], []
            if agente and agente != "Todos":
                conditions.append('"Nome do Agente" = ?')
                params.append(agente)
            if equipe and equipe != "Todas":
                conditions.append('Equipe = ?')
                params.append(equipe)
            if avaliacao:
                conditions.append('"Avaliação ATD." = ?')
                params.append(str(avaliacao))
            if pontuacao:
                conditions.append('Pontuação = ?')
                params.append(str(pontuacao))
            expr_ymd = "substr(\"Data M\",7,4) || substr(\"Data M\",4,2) || substr(\"Data M\",1,2)"
            if data_ini:
                conditions.append(f"{expr_ymd} >= ?")
                params.append(data_ini.strftime('%Y%m%d'))
            if data_fim:
                conditions.append(f"{expr_ymd} <= ?")
                params.append(data_fim.strftime('%Y%m%d'))
            if conditions:
                query += " WHERE " + " AND ".join(conditions)
            df_lancamentos = pd.read_sql_query(query, conn, params=params)
        if 'id' in df_lancamentos.columns:
            df_lancamentos = df_lancamentos.drop(columns=['id'])
        # Garante a ordem e a presença de todas as colunas
        df_lancamentos = df_lancamentos.reindex(columns=COLUNAS).fillna('')

        # Calcular métricas gerais
        total_monitorias = pd.to_numeric(df_dashboard['Total Monitorias'], errors='coerce').sum()
        media_pontuacao = df_dashboard['Média Pontuação'].mean()
        media_erro_critico = pd.to_numeric(df_dashboard['Média Erro Crítico (%)'], errors='coerce').mean()

        resumo_data = {
            'Métrica': ['Total de Monitorias', 'Média Geral de Pontuação', 'Média de Erros Críticos por Agente (%)'],
            'Valor': [f"{total_monitorias:.0f}", f"{media_pontuacao:.2f}", f"{media_erro_critico:.2f}"]
        }
        df_resumo = pd.DataFrame(resumo_data)

        # Salvar gráficos temporariamente
        temp_dir = os.getcwd()
        bar_path = os.path.join(temp_dir, f"temp_bar_{timestamp}.png")
        pie_path = os.path.join(temp_dir, f"temp_pie_{timestamp}.png")

        # Gráfico de barras (com dados do dashboard filtrado)
        fig_bar, ax_bar = plt.subplots(figsize=(8, 5))
        ax_bar.bar(df_dashboard['Agente'], df_dashboard['Média Pontuação'], color='#4A90E2')
        ax_bar.set_title('Média de Pontuação por Agente')
        ax_bar.set_xlabel('Agente')
        ax_bar.set_ylabel('Pontuação')
        ax_bar.set_ylim(0, 10)
        ax_bar.tick_params(axis='x', rotation=45)
        plt.tight_layout()
        fig_bar.savefig(bar_path)
        plt.close(fig_bar)

        # Gráfico de pizza (com dados do dashboard filtrado)
        total_erros_criticos = pd.to_numeric(df_dashboard['Erros Críticos'], errors='coerce').sum()
        total_sem_erros = total_monitorias - total_erros_criticos
        
        if total_monitorias > 0:
            labels = ['Com Erro Crítico', 'Sem Erro Crítico']
            sizes = [total_erros_criticos, total_sem_erros]
            colors = ['#FF4C4C', '#4A90E2']
            explode = [0.1, 0]

            fig_pie, ax_pie = plt.subplots(figsize=(5, 5))
            ax_pie.pie(sizes, explode=explode, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
            ax_pie.set_title('Proporção de Erros Críticos')
            plt.tight_layout()
            fig_pie.savefig(pie_path)
            plt.close(fig_pie)
        
        # Escrever no Excel
        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            df_resumo.to_excel(writer, sheet_name='Resumo', index=False)
            df_dashboard.to_excel(writer, sheet_name='Dashboard', index=False)
            
            # CORREÇÃO: Converte 'Pontuação' para numérico antes de comparar
            df_lancamentos['Pontuação'] = pd.to_numeric(df_lancamentos['Pontuação'], errors='coerce')
            ranking_zero = df_lancamentos[df_lancamentos['Pontuação'] == 0]['Monitoria Zero'].value_counts().reset_index()
            ranking_zero.columns = ['Motivo', 'Quantidade']
            ranking_zero['Motivo'] = ranking_zero['Motivo'].replace('', 'Não especificado').astype(str)
            ranking_zero.to_excel(writer, sheet_name='Ranking Zeros', index=False)

            df_lancamentos['Pontuação'] = df_lancamentos['Pontuação'].apply(lambda x: f"{x:.2f}") # Reverte para string para exibição
            df_lancamentos.to_excel(writer, sheet_name='Lançamentos Completos', index=False)

        # Adicionar filtros aplicados ao Excel (aba Resumo)
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image
        wb = load_workbook(output_excel)
        ws_resumo = wb['Resumo']
        try:
            periodo_texto = ""
            if entry_data_ini_dashboard.get() and entry_data_fim_dashboard.get():
                periodo_texto = f"Período: {entry_data_ini_dashboard.get()} a {entry_data_fim_dashboard.get()}"
            elif entry_data_ini_dashboard.get():
                periodo_texto = f"Período: a partir de {entry_data_ini_dashboard.get()}"
            elif entry_data_fim_dashboard.get():
                periodo_texto = f"Período: até {entry_data_fim_dashboard.get()}"
            if periodo_texto:
                ws_resumo.cell(row=5, column=1, value=periodo_texto)
        except Exception:
            pass
        
        img_bar = Image(bar_path)
        img_bar.width, img_bar.height = 600, 375
        ws_resumo.add_image(img_bar, 'A10')

        if total_monitorias > 0 and os.path.exists(pie_path):
            img_pie = Image(pie_path)
            img_pie.width, img_pie.height = 375, 375
            ws_resumo.add_image(img_pie, 'J10')
        
        wb.save(output_excel)

        # Remover arquivos temporários
        for file in [bar_path, pie_path]:
            if os.path.exists(file): os.remove(file)

        messagebox.showinfo("Sucesso", f"Relatório gerado com sucesso em: {output_excel}")

    except Exception as e:
        messagebox.showerror("Erro ao Gerar Relatório", f"Ocorreu um erro: {e}\n\nVerifique se o arquivo Excel não está aberto.")
        # Limpa arquivos temporários em caso de falha
        for file in [bar_path, pie_path]:
            if 'path' in locals() and os.path.exists(file): os.remove(file)


def editar_registro():
    """Carrega o registro selecionado para edição no formulário."""
    global edit_mode, edit_id
    selected_items = tree.selection()
    if not selected_items:
        messagebox.showwarning("Nenhum Registro Selecionado", "Selecione um registro para editar.")
        return
    
    item = selected_items[0]
    values = tree.item(item, 'values')
    try:
        edit_id = int(item)
    except ValueError:
        protocolo_selecionado = values[COLUNAS.index('Protocolo')]
        try:
            with sqlite3.connect(DB_FILE) as conn:
                result = conn.execute("SELECT id FROM monitoria WHERE Protocolo = ?", (protocolo_selecionado,)).fetchone()
                if not result:
                    messagebox.showerror("Erro", "Não foi possível encontrar o ID do registro.")
                    return
                edit_id = result[0]
        except Exception as e:
            messagebox.showerror("Erro de Banco de Dados", f"Não foi possível buscar o ID do registro: {e}")
            return

    limpar_formulario() # Limpa o formulário antes de preencher
    for i, col in enumerate(COLUNAS):
        value = values[i] if i < len(values) else ""
        if col in widgets and widgets[col]:
            if col in YES_NO_FIELDS or col == 'Erro Crítico?' or col in ['Nome do Agente', 'Equipe']:
                widgets[col].set(value)
                if col in CRITICAL_ERRORS:
                    atualizar_cor_critica(widgets[col], col)
            elif col == 'Data M':
                try:
                    widgets[col].set_date(datetime.strptime(value, '%d/%m/%Y'))
                except (ValueError, TypeError):
                    widgets[col].set_date(datetime.now())
            elif col in ['Protocolo', 'Avaliação ATD.']:
                widgets[col].insert(0, value)
            elif col == 'Observações':
                widgets[col].insert("1.0", value)
            elif col == 'Motivo do Atendimento':
                widgets[col].set(value)
            elif col == 'Monitoria Zero':
                widgets[col].set(value if value else 'Nenhum')

    edit_mode = True
    botao_salvar.configure(text="Atualizar Monitoria")
    tabview.set("Nova Monitoria")

def analisar_protocolo_com_ia():
    """
    Orquestra a busca do chat, análise pela IA e preenchimento do formulário.
    """
    protocolo = widgets['Protocolo'].get()
    if not protocolo:
        messagebox.showwarning("Protocolo Vazio", "Por favor, insira um número de protocolo para analisar.")
        return

    # 1. Exibe uma mensagem de "analisando"
    messagebox.showinfo("Analisando", f"Buscando e analisando o protocolo {protocolo} com a IA. Por favor, aguarde...")
    app.update_idletasks() # Força a atualização da UI para mostrar o popup

    try:
        # 2. Busca o histórico do chat
        transcript = analyzer.fetch_chat_history(protocolo)
        if transcript.startswith("ERRO:"):
            messagebox.showerror("Erro na API", transcript)
            return

        # 3. Envia para análise da IA
        # Passa apenas os campos que a IA deve avaliar
        analysis_result = analyzer.analyze_transcript_with_gemini(transcript, YES_NO_FIELDS)

        if "error" in analysis_result:
            messagebox.showerror("Erro na IA", analysis_result["error"])
            if "Observações" in analysis_result:
                 widgets['Observações'].delete("1.0", tk.END)
                 widgets['Observações'].insert("1.0", analysis_result["Observações"])
            return

        # 4. Preenche o formulário com os resultados
        for field, value in analysis_result.items():
            if field in widgets and widgets[field]:
                if field == 'Observações':
                    widgets[field].delete("1.0", tk.END)
                    widgets[field].insert("1.0", value)
                else:
                    # Garante que o valor retornado pela IA está entre as opções do ComboBox
                    if value in widgets[field].cget('values'):
                        widgets[field].set(value)
        messagebox.showinfo("Sucesso", "Análise da IA concluída e formulário preenchido!")
    except Exception as e:
        messagebox.showerror("Erro Inesperado", f"Ocorreu um erro durante a análise: {e}")

def _salvar_dados_auditoria(dados_ia: dict):
    """
    Função interna para salvar os dados da auditoria da IA no banco de dados.
    """
    pontuacao, itens_aplicaveis, erro_critico = calcular_pontuacao(dados_ia)
    dados_ia['Pontuação'] = f"{pontuacao:.2f}"
    dados_ia['Itens Aplicáveis'] = str(itens_aplicaveis)
    dados_ia['Erro Crítico?'] = erro_critico

    for col in COLUNAS:
        if col not in dados_ia:
            dados_ia[col] = ''
    
    if verificar_protocolo_duplicado(dados_ia['Protocolo']):
        print(f"Protocolo {dados_ia['Protocolo']} já existe no banco. Pulando.")
        return False

    try:
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            columns = ', '.join([f'"{col}"' for col in COLUNAS])
            placeholders = ', '.join(['?' for _ in COLUNAS])
            values = [dados_ia.get(col, '') for col in COLUNAS]
            cursor.execute(f'INSERT INTO monitoria ({columns}) VALUES ({placeholders})', values)
            conn.commit()
        return True
    except Exception as e:
        print(f"Erro ao salvar auditoria para o protocolo {dados_ia.get('Protocolo', 'N/A')}: {e}")
        return False

def auditar_periodo_com_ia():
    """
    Orquestra o processo de auditoria em massa por período e departamentos.
    """
    data_ini = entry_data_ini_dashboard.get_date() if entry_data_ini_dashboard.get() else None
    data_fim = entry_data_fim_dashboard.get_date() if entry_data_fim_dashboard.get() else None

    if not data_ini or not data_fim:
        messagebox.showerror("Filtros Incompletos", "Por favor, selecione uma data de início e fim.")
        return

    # --- NOVA LÓGICA: Coleta os departamentos selecionados ---
    deptos_selecionados = [depto for depto, var in check_dept_vars.items() if var.get() == 1]
    
    if not deptos_selecionados:
        messagebox.showerror("Filtros Incompletos", "Por favor, selecione pelo menos um departamento para auditar.")
        return

    deptos_str = ", ".join(deptos_selecionados)
    if not messagebox.askyesno("Confirmar Auditoria em Massa", 
                               f"Deseja iniciar a auditoria por IA para:\n\n"
                               f"Período: {data_ini.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}\n"
                               f"Departamentos: {deptos_str}\n\n"
                               "Este processo pode levar vários minutos."):
        return

    messagebox.showinfo("Buscando Atendimentos", "Buscando atendimentos na API. Aguarde...")
    app.update_idletasks()
    
    try:
        # Passa a lista de departamentos para a função do analyzer
        atendimentos_para_auditar = analyzer.fetch_attendances_by_date_range(data_ini, data_fim, deptos_selecionados)
    except analyzer.APIError as e:
        messagebox.showerror("Erro na API", f"Não foi possível buscar os atendimentos:\n\n{e}")
        return

    if not atendimentos_para_auditar:
        messagebox.showinfo("Nenhum Atendimento", "Nenhum atendimento de chat encontrado para os filtros selecionados.")
        return
    
    total_atendimentos = len(atendimentos_para_auditar)
    
    # (O restante da função continua igual: barra de progresso, loop, salvamento, etc.)
    progress_popup = ctk.CTkToplevel(app)
    progress_popup.title("Auditando...")
    progress_popup.geometry("400x100")
    progress_popup.transient(app)
    progress_popup.attributes('-topmost', True)

    ctk.CTkLabel(progress_popup, text="Analisando atendimentos com a IA...").pack(pady=10)
    progress_bar = ctk.CTkProgressBar(progress_popup, width=350)
    progress_bar.set(0)
    progress_bar.pack(pady=5)
    progress_label = ctk.CTkLabel(progress_popup, text=f"0 de {total_atendimentos}")
    progress_label.pack()

    salvos_com_sucesso = 0
    for i, atendimento in enumerate(atendimentos_para_auditar):
        try:
            progresso_atual = (i + 1) / total_atendimentos
            progress_bar.set(progresso_atual)
            progress_label.configure(text=f"{i + 1} de {total_atendimentos} (Protocolo: {atendimento['protocolo']})")
            app.update_idletasks()

            analysis_result = analyzer.analyze_transcript_with_gemini(atendimento['transcript'], YES_NO_FIELDS)
            
            if "error" in analysis_result:
                print(f"Erro da IA no protocolo {atendimento['protocolo']}: {analysis_result['error']}")
                continue

            dados_para_salvar = analysis_result.copy()
            dados_para_salvar['Protocolo'] = atendimento['protocolo']
            
            try:
                data_atd_obj = datetime.fromisoformat(atendimento['dataAtendimento'].replace('Z', '+00:00'))
                dados_para_salvar['Data M'] = data_atd_obj.strftime('%d/%m/%Y')
            except:
                dados_para_salvar['Data M'] = datetime.now().strftime('%d/%m/%Y')

            nome_agente_api = atendimento.get('nomeAgente')
            if nome_agente_api:
                dados_para_salvar['Nome do Agente'] = nome_agente_api
                with sqlite3.connect(DB_FILE) as conn:
                    cursor = conn.cursor()
                    cursor.execute('SELECT equipe FROM agentes WHERE nome = ?', (nome_agente_api,))
                    row = cursor.fetchone()
                    dados_para_salvar['Equipe'] = row[0] if row else 'Equipe Desconhecida'
            
            if _salvar_dados_auditoria(dados_para_salvar):
                salvos_com_sucesso += 1
        
        except Exception as e:
            print(f"Erro crítico no loop de auditoria para o protocolo {atendimento['protocolo']}: {e}")

    progress_popup.destroy()
    update_excel()
    aplicar_filtros()
    aplicar_filtros_dashboard()
    
    messagebox.showinfo("Auditoria Concluída", 
                        f"Processo finalizado!\n\n"
                        f"Total de atendimentos encontrados: {total_atendimentos}\n"
                        f"Auditados e salvos com sucesso: {salvos_com_sucesso}\n\n"
                        "O dashboard e o arquivo Excel foram atualizados.")

# --- CONFIGURAÇÃO DA JANELA PRINCIPAL ---
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

app = ctk.CTk()
app.title("Sistema de Monitoria")

# Define ícone da aplicação, se disponível
app_icon_img = None
try:
    if os.path.exists(APP_ICON_FILE):
        app_icon_img = tk.PhotoImage(file=APP_ICON_FILE)
        app.iconphoto(True, app_icon_img)
except Exception:
    pass

screen_width, screen_height = app.winfo_screenwidth(), app.winfo_screenheight()
window_width, window_height = 1200, 750
x, y = (screen_width - window_width) // 2, (screen_height - window_height) // 2
app.geometry(f"{window_width}x{window_height}+{x}+{y}")
app.minsize(1100, 700)

# Exibe a logo no topo, se disponível; caso contrário, mostra texto padrão
label_canaa = None
try:
    if os.path.exists(APP_LOGO_FILE):
        pil_img = Image.open(APP_LOGO_FILE)
        # Redimensiona preservando proporção para caber bem no topo
        max_w, max_h = 300, 60
        w, h = pil_img.size
        scale = min(max_w / float(w), max_h / float(h))
        new_size = (max(1, int(w * scale)), max(1, int(h * scale)))
        ctk_img = ctk.CTkImage(light_image=pil_img.resize(new_size, Image.LANCZOS), size=new_size)
        app_logo_img = ctk_img  # manter referência global
        label_canaa = ctk.CTkLabel(app, image=ctk_img, text="")
        label_canaa.pack(pady=(10, 5))
    else:
        label_canaa = ctk.CTkLabel(app, text="Canaã Telecom", font=ctk.CTkFont(family="Helvetica", size=28, weight="bold"), text_color="#FFFFFF")
        label_canaa.pack(pady=(10, 5))
except Exception:
    label_canaa = ctk.CTkLabel(app, text="Canaã Telecom", font=ctk.CTkFont(family="Helvetica", size=28, weight="bold"), text_color="#FFFFFF")
    label_canaa.pack(pady=(10, 5))

tabview = ctk.CTkTabview(app, fg_color="#1C2526", segmented_button_fg_color="#4A90E2", segmented_button_selected_color="#2E5A88", text_color="#FFFFFF")
tabview.pack(pady=5, padx=10, fill="both", expand=True)

tab_form = tabview.add("Nova Monitoria")
tab_table = tabview.add("Últimos Lançamentos")
tab_dashboard = tabview.add("Dashboard")
tab_agentes = tabview.add("Configurações")

# --- FORMULÁRIO ---
# Garante que o banco esteja inicializado antes de carregar comboboxes
init_db()
form_frame = ctk.CTkScrollableFrame(tab_form, fg_color="transparent")
form_frame.pack(pady=10, padx=10, fill="both", expand=True)

num_columns = 5
for i in range(num_columns):
    form_frame.grid_columnconfigure(i, weight=1)

# Fontes padrão para a aba 'Nova Monitoria' (Poppins)
FONT_POPPINS_12 = ctk.CTkFont(family="Poppins", size=12)
FONT_POPPINS_14_BOLD = ctk.CTkFont(family="Poppins", size=14, weight="bold")
FONT_POPPINS_20_BOLD = ctk.CTkFont(family="Poppins", size=20, weight="bold")

label_titulo = ctk.CTkLabel(form_frame, text="Nova Monitoria", font=FONT_POPPINS_20_BOLD, text_color="#FFFFFF")
label_titulo.grid(row=0, column=0, columnspan=num_columns, pady=(5, 15), sticky="n")

widgets = {col: None for col in COLUNAS}
equipes, agentes = carregar_dados_iniciais()

def create_widget_frame(parent, label_text, widget_class, widget_kwargs, row, col, colspan=1):
    frame = ctk.CTkFrame(parent, fg_color="transparent")
    frame.grid(row=row, column=col, columnspan=colspan, padx=5, pady=(2, 8), sticky="ew")
    
    text_color = "#FF0000" if label_text in CRITICAL_ERRORS else "#FFFFFF"
    label = ctk.CTkLabel(frame, text=label_text, font=FONT_POPPINS_12, text_color=text_color)
    label.pack(anchor="w", padx=5)

    if widget_class == DateEntry:
        widget = widget_class(frame, **widget_kwargs)
    else:
        widget = widget_class(frame, **widget_kwargs)

    widget.pack(fill="x", expand=True, padx=5)
    # Tenta aplicar fonte padrão Poppins aos widgets suportados
    try:
        widget.configure(font=FONT_POPPINS_12)
    except Exception:
        pass
    return widget

main_fields_layout = [
    ('Protocolo', ctk.CTkEntry, {'fg_color': '#2E5A88', 'text_color': '#FFFFFF', 'font': FONT_POPPINS_12}),
    ('Data M', DateEntry, {'width': 14, 'background': '#4A90E2', 'foreground': 'white', 'borderwidth': 2, 'date_pattern': 'dd/mm/yyyy'}),
    ('Nome do Agente', ctk.CTkComboBox, {'values': agentes, 'fg_color': '#2E5A88', 'text_color': '#FFFFFF', 'font': FONT_POPPINS_12}),
    ('Equipe', ctk.CTkComboBox, {'values': equipes, 'fg_color': '#2E5A88', 'text_color': '#FFFFFF', 'font': FONT_POPPINS_12}),
]
current_row = 1
for idx, (col, widget_type, kwargs) in enumerate(main_fields_layout):
    widgets[col] = create_widget_frame(form_frame, col, widget_type, kwargs, current_row, idx)
widgets['Data M'].set_date(datetime.now())
widgets['Nome do Agente'].set('')
widgets['Equipe'].set('')
form_frame.grid_rowconfigure(current_row, minsize=80)
widgets['Nome do Agente'].bind('<<ComboboxSelected>>', atualizar_equipe)
current_row += 1

for idx, col in enumerate(YES_NO_FIELDS):
    col_idx = idx % num_columns
    row_idx = current_row + (idx // num_columns)
    form_frame.grid_rowconfigure(row_idx, minsize=80)
    widgets[col] = create_widget_frame(form_frame, col, ctk.CTkComboBox, {'values': ['Conforme', 'Não Conforme', 'Não se aplica'], 'fg_color': '#2E5A88', 'text_color': '#FFFFFF', 'font': FONT_POPPINS_12}, row_idx, col_idx)
    widgets[col].set('Conforme')
    if col in CRITICAL_ERRORS:
        widgets[col].bind('<<ComboboxSelected>>', lambda e, w=widgets[col], c=col: atualizar_cor_critica(w, c))
current_row += (len(YES_NO_FIELDS) + num_columns - 1) // num_columns

form_frame.grid_rowconfigure(current_row, minsize=80)
widgets['Motivo do Atendimento'] = create_widget_frame(form_frame, 'Motivo do Atendimento', ctk.CTkComboBox, {'values': ['Conexão', 'Cancelamento', 'Reclamação', 'Financeiro', 'Outros Assuntos'], 'fg_color': '#2E5A88', 'text_color': '#FFFFFF', 'font': FONT_POPPINS_12}, current_row, 0)
widgets['Monitoria Zero'] = create_widget_frame(form_frame, 'Monitoria Zero', ctk.CTkComboBox, {'values': ['Nenhum', 'Ofensa Verbal', 'Erro de Procedimento', 'Confirmação de dados', 'Omissão de Atendimento', 'Inf. Protocolo'], 'fg_color': '#2E5A88', 'text_color': '#FFFFFF', 'font': FONT_POPPINS_12}, current_row, 1)
widgets['Motivo do Atendimento'].set('')
widgets['Monitoria Zero'].set('Nenhum')
current_row += 1

other_fields_layout = [
    ('Avaliação ATD.', ctk.CTkEntry, {'fg_color': '#2E5A88', 'text_color': '#FFFFFF', 'font': FONT_POPPINS_12}),
    ('Erro Crítico?', ctk.CTkComboBox, {'values': ['Sim', 'Não'], 'fg_color': '#2E5A88', 'text_color': '#FFFFFF', 'font': FONT_POPPINS_12}),
    ('Inf. Protocolo?', ctk.CTkComboBox, {'values': ['Conforme', 'Não Conforme', 'Não se aplica'], 'fg_color': '#2E5A88', 'text_color': '#FFFFFF', 'font': FONT_POPPINS_12}),
]
form_frame.grid_rowconfigure(current_row, minsize=80)
for idx, (col, widget_type, kwargs) in enumerate(other_fields_layout):
    widgets[col] = create_widget_frame(form_frame, col, widget_type, kwargs, current_row, idx)
widgets['Erro Crítico?'].set('Não')
widgets['Inf. Protocolo?'].set('Conforme')
current_row += 1

form_frame.grid_rowconfigure(current_row, minsize=100)
widgets['Observações'] = create_widget_frame(form_frame, 'Observações', ctk.CTkTextbox, {'height': 100, 'fg_color': '#2E5A88', 'text_color': '#FFFFFF', 'font': FONT_POPPINS_12}, current_row, 0, colspan=num_columns)
current_row += 1

button_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
button_frame.grid(row=current_row, column=0, columnspan=num_columns, pady=20)
botao_salvar = ctk.CTkButton(button_frame, text="Salvar Monitoria", command=salvar_monitoria, font=ctk.CTkFont(family="Arial", size=14, weight="bold"), fg_color="#4A90E2", hover_color="#2E5A88")
botao_salvar.pack()

# Botão para análise com IA
botao_analise_ia = ctk.CTkButton(form_frame, text="Analisar Protocolo com IA 🤖", command=analisar_protocolo_com_ia, fg_color="#17A2B8", hover_color="#138496")
botao_analise_ia.grid(row=1, column=4, padx=5, pady=5, sticky="ew")

# --- TABELA ---
table_frame = ctk.CTkFrame(tab_table, fg_color="transparent")
table_frame.pack(pady=10, padx=10, fill="both", expand=True)

filter_frame = ctk.CTkFrame(table_frame, fg_color="transparent")
filter_frame.pack(pady=5, padx=10, fill="x")

ctk.CTkLabel(filter_frame, text="Filtrar por Agente:", font=ctk.CTkFont(family="Arial", size=12)).pack(side="left", padx=5)
combo_filtro_agente = ctk.CTkComboBox(filter_frame, values=["Todos"] + agentes, width=200, fg_color="#2E5A88", text_color="#FFFFFF")
combo_filtro_agente.pack(side="left", padx=5)
combo_filtro_agente.set("Todos")

ctk.CTkLabel(filter_frame, text="Filtrar por Protocolo:", font=ctk.CTkFont(family="Arial", size=12)).pack(side="left", padx=5)
entry_filtro_protocolo = ctk.CTkEntry(filter_frame, width=150, fg_color="#2E5A88", text_color="#FFFFFF")
entry_filtro_protocolo.pack(side="left", padx=5)

ctk.CTkButton(filter_frame, text="Filtrar", command=aplicar_filtros, font=ctk.CTkFont(family="Arial", size=14, weight="bold"), fg_color="#4A90E2", hover_color="#2E5A88").pack(side="left", padx=5)
ctk.CTkButton(filter_frame, text="Limpar Filtros", command=limpar_filtros, font=ctk.CTkFont(family="Arial", size=14, weight="bold"), fg_color="#4A90E2", hover_color="#2E5A88").pack(side="left", padx=5)

ctk.CTkLabel(table_frame, text="Últimos Lançamentos", font=ctk.CTkFont(family="Arial", size=16, weight="bold")).pack(pady=(5, 10))

style = ttk.Style()
style.theme_use("default")
style.configure("Treeview", background="#2a2d2e", foreground="#FFFFFF", fieldbackground="#2a2d2e", borderwidth=0, rowheight=25)
style.map('Treeview', background=[('selected', '#4A90E2')])
style.configure("Treeview.Heading", background="#2E5A88", foreground="#FFFFFF", font=('Arial', 10, 'bold'), relief="flat")
style.map("Treeview.Heading", background=[('active', '#3B6EA8')])

tree_container = ctk.CTkFrame(table_frame, fg_color="transparent")
tree_container.pack(fill="both", expand=True)

tree = ttk.Treeview(tree_container, columns=COLUNAS, show='headings')
vsb = ctk.CTkScrollbar(tree_container, orientation="vertical", command=tree.yview)
vsb.pack(side='right', fill='y')
hsb = ctk.CTkScrollbar(tree_container, orientation="horizontal", command=tree.xview)
hsb.pack(side='bottom', fill='x')
tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
tree.pack(fill="both", expand=True)

# Botão flutuante de chat (FAB)
def toggle_chat_window():
    global chat_window
    if chat_window and tk.Toplevel.winfo_exists(chat_window):
        chat_window.destroy()
        chat_window = None
        return
    chat_window = tk.Toplevel(app)
    chat_window.title("Assistente (pré-IA)")
    chat_window.geometry("360x480")
    chat_window.transient(app)
    chat_window.attributes('-topmost', True)
    # Área de mensagens
    chat_display = tk.Text(chat_window, state='disabled', wrap='word', bg='#1E1E1E', fg='#FFFFFF')
    chat_display.pack(fill='both', expand=True, padx=8, pady=(8,4))
    # Área de envio
    entry_frame = ctk.CTkFrame(chat_window)
    entry_frame.pack(fill='x', padx=8, pady=(0,8))
    chat_entry = ctk.CTkEntry(entry_frame, placeholder_text='Digite sua mensagem...')
    chat_entry.pack(side='left', fill='x', expand=True, padx=(0,8))
    send_btn = ctk.CTkButton(entry_frame, text='Enviar', command=lambda: None)
    send_btn.pack(side='left')
    # Notas: integração de IA será adicionada futuramente; UI já preparada.

# Cria um FAB no canto inferior direito do app
fab_frame = ctk.CTkFrame(app, corner_radius=30, fg_color='transparent')
fab_frame.place(relx=1.0, rely=1.0, x=-20, y=-20, anchor='se')
chat_fab = ctk.CTkButton(fab_frame, text="💬", width=50, height=50, corner_radius=25, font=ctk.CTkFont(size=20, weight='bold'),
                        command=toggle_chat_window, fg_color="#4A90E2", hover_color="#2E5A88")
chat_fab.pack()

button_table_frame = ctk.CTkFrame(table_frame, fg_color="transparent")
button_table_frame.pack(pady=(10, 5))
ctk.CTkButton(button_table_frame, text="Editar Selecionado", command=editar_registro, font=ctk.CTkFont(family="Arial", size=14, weight="bold"), fg_color="#4A90E2", hover_color="#2E5A88").pack(side="left", padx=5)
ctk.CTkButton(button_table_frame, text="Excluir Selecionado", command=excluir_registro, font=ctk.CTkFont(family="Arial", size=14, weight="bold"), fg_color="#FF4C4C", hover_color="#CC3333").pack(side="left", padx=5)

# --- DASHBOARD ---
dashboard_frame = ctk.CTkScrollableFrame(tab_dashboard, fg_color="transparent")
dashboard_frame.pack(pady=10, padx=10, fill="both", expand=True)

filter_dashboard_frame = ctk.CTkFrame(dashboard_frame, fg_color="transparent")
filter_dashboard_frame.pack(pady=5, padx=10, fill="x")
depto_filter_frame = ctk.CTkFrame(dashboard_frame, fg_color="transparent")
depto_filter_frame.pack(pady=(5, 10), padx=10, fill="x")

ctk.CTkLabel(filter_dashboard_frame, text="Agente:").pack(side="left", padx=(5,2))
combo_filtro_agente_dashboard = ctk.CTkComboBox(filter_dashboard_frame, values=["Todos"] + agentes, width=180, fg_color="#2E5A88")
combo_filtro_agente_dashboard.pack(side="left", padx=5)
combo_filtro_agente_dashboard.set("Todos")

ctk.CTkLabel(filter_dashboard_frame, text="Equipe:").pack(side="left", padx=(5,2))
combo_filtro_equipe_dashboard = ctk.CTkComboBox(filter_dashboard_frame, values=["Todas"] + equipes, width=140, fg_color="#2E5A88")
combo_filtro_equipe_dashboard.pack(side="left", padx=5)
combo_filtro_equipe_dashboard.set("Todas")

ctk.CTkLabel(filter_dashboard_frame, text="Avaliação ATD.:").pack(side="left", padx=(5,2))
entry_filtro_avaliacao_dashboard = ctk.CTkEntry(filter_dashboard_frame, width=80, fg_color="#2E5A88")
entry_filtro_avaliacao_dashboard.pack(side="left", padx=5)

ctk.CTkLabel(filter_dashboard_frame, text="Pontuação:").pack(side="left", padx=(5,2))
entry_filtro_pontuacao_dashboard = ctk.CTkEntry(filter_dashboard_frame, width=80, fg_color="#2E5A88")
entry_filtro_pontuacao_dashboard.pack(side="left", padx=5)

# Filtros de período (Data M)
ctk.CTkLabel(filter_dashboard_frame, text="De:").pack(side="left", padx=(10,2))
entry_data_ini_dashboard = DateEntry(filter_dashboard_frame, width=12, background="#4A90E2", foreground="white", borderwidth=2, date_pattern='dd/mm/yyyy')
entry_data_ini_dashboard.pack(side="left", padx=5)
entry_data_ini_dashboard.set_date(datetime.now().replace(day=1))

ctk.CTkLabel(filter_dashboard_frame, text="Até:").pack(side="left", padx=(5,2))
entry_data_fim_dashboard = DateEntry(filter_dashboard_frame, width=12, background="#4A90E2", foreground="white", borderwidth=2, date_pattern='dd/mm/yyyy')
entry_data_fim_dashboard.pack(side="left", padx=5)
entry_data_fim_dashboard.set_date(datetime.now())

ctk.CTkButton(filter_dashboard_frame, text="Filtrar", command=aplicar_filtros_dashboard, font=ctk.CTkFont(weight="bold")).pack(side="left", padx=5)
ctk.CTkButton(filter_dashboard_frame, text="Limpar", command=limpar_filtros_dashboard, font=ctk.CTkFont(weight="bold")).pack(side="left", padx=5)

ctk.CTkLabel(dashboard_frame, text="Dashboard", font=ctk.CTkFont(size=20, weight="bold")).pack(pady=(5, 10))

dashboard_container = ctk.CTkFrame(dashboard_frame, fg_color="transparent")
dashboard_container.pack(fill="both", expand=True)
dashboard_tree = ttk.Treeview(dashboard_container, show='headings')
dashboard_vsb = ctk.CTkScrollbar(dashboard_container, orientation="vertical", command=dashboard_tree.yview)
dashboard_vsb.pack(side='right', fill='y')
dashboard_hsb = ctk.CTkScrollbar(dashboard_container, orientation="horizontal", command=dashboard_tree.xview)
dashboard_hsb.pack(side='bottom', fill='x')
dashboard_tree.configure(yscrollcommand=dashboard_vsb.set, xscrollcommand=dashboard_hsb.set)
dashboard_tree.pack(fill="both", expand=True)

charts_frame = ctk.CTkFrame(dashboard_frame, fg_color="transparent")
charts_frame.pack(pady=10, fill="x", expand=True)

button_dashboard_frame = ctk.CTkFrame(dashboard_frame, fg_color="transparent")
button_dashboard_frame.pack(pady=10, fill="x")

ctk.CTkButton(
    button_dashboard_frame,
    text="🤖 Auditar Período com IA",
    command=auditar_periodo_com_ia,
    font=ctk.CTkFont(size=16, weight="bold"),
    fg_color="#007BFF",
    hover_color="#0056b3",
    height=40
).pack(pady=10)

ctk.CTkButton(button_dashboard_frame, text="Exportar Relatório", command=gerar_relatorio, font=ctk.CTkFont(size=16, weight="bold"), fg_color="#28A745", hover_color="#218838", height=40).pack(pady=10)

# --- CONFIGURAÇÕES ---
agentes_frame = ctk.CTkFrame(tab_agentes, fg_color="transparent")
agentes_frame.pack(pady=10, padx=10, fill="both", expand=True)

gerenciar_agentes_frame = ctk.CTkFrame(agentes_frame, fg_color="transparent")
gerenciar_agentes_frame.pack(pady=10, padx=10, fill="x")
ctk.CTkLabel(gerenciar_agentes_frame, text="Gerenciar Agentes", font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(5, 10))

ctk.CTkLabel(agentes_frame, text="Configurações", font=ctk.CTkFont(size=20, weight="bold")).pack(pady=(5, 15))

adicionar_frame = ctk.CTkFrame(agentes_frame, fg_color="transparent")
adicionar_frame.pack(pady=10, fill="x")

ctk.CTkLabel(adicionar_frame, text="Nome do Agente:").pack(side="left", padx=5)
entry_novo_agente = ctk.CTkEntry(adicionar_frame, width=200, fg_color="#2E5A88")
entry_novo_agente.pack(side="left", padx=5)

ctk.CTkLabel(adicionar_frame, text="Equipe:").pack(side="left", padx=5)
combo_equipe_novo_agente = ctk.CTkComboBox(adicionar_frame, values=equipes, width=150, fg_color="#2E5A88")
combo_equipe_novo_agente.pack(side="left", padx=5)
combo_equipe_novo_agente.set('')

botao_adicionar_agente = ctk.CTkButton(adicionar_frame, text="Adicionar Agente", command=adicionar_agente, font=ctk.CTkFont(weight="bold"))
botao_adicionar_agente.pack(side="left", padx=10)
ctk.CTkButton(adicionar_frame, text="Alterar Senha Admin", command=alterar_senha_admin, font=ctk.CTkFont(weight="bold")).pack(side="left", padx=10)

list_frame = ctk.CTkFrame(gerenciar_agentes_frame, fg_color="transparent")
list_frame.pack(pady=10, fill="both", expand=True)
ctk.CTkLabel(list_frame, text="Agentes Cadastrados:", anchor="w").pack(fill="x", padx=5)
listbox_agentes = tk.Listbox(list_frame, height=10, bg="#2a2d2e", fg="#FFFFFF", font=("Arial", 12), selectbackground="#4A90E2", selectforeground="#FFFFFF")
listbox_agentes.pack(fill="both", expand=True, padx=5, pady=5)
listbox_scroll = ctk.CTkScrollbar(list_frame, orientation="vertical", command=listbox_agentes.yview)
listbox_scroll.pack(side="right", fill="y")
listbox_agentes.configure(yscrollcommand=listbox_scroll.set)
try:
    with sqlite3.connect(DB_FILE) as conn:
        df_init = pd.read_sql_query('SELECT nome, equipe FROM agentes ORDER BY nome COLLATE NOCASE', conn)
    for _, r in df_init.iterrows():
        listbox_agentes.insert(tk.END, f"{r['nome']} ({r['equipe']})")
except Exception:
    for agente in sorted(AGENTES_EQUIPE.keys()):
        listbox_agentes.insert(tk.END, f"{agente} ({AGENTES_EQUIPE[agente]})")

button_agentes_frame = ctk.CTkFrame(gerenciar_agentes_frame, fg_color="transparent")
button_agentes_frame.pack(pady=5)
ctk.CTkButton(button_agentes_frame, text="Excluir Agente", command=excluir_agente, font=ctk.CTkFont(weight="bold"), fg_color="#FF4C4C", hover_color="#CC3333").pack(side="left", padx=5)
ctk.CTkButton(button_agentes_frame, text="Editar Agente", command=editar_agente, font=ctk.CTkFont(weight="bold")).pack(side="left", padx=5)
ctk.CTkButton(button_agentes_frame, text="Limpar Lançamentos", command=limpar_lancamentos, font=ctk.CTkFont(weight="bold"), fg_color="#FF4C4C", hover_color="#CC3333").pack(side="left", padx=5)

# --- Separador ---
separator = ttk.Separator(agentes_frame, orient='horizontal')
separator.pack(fill='x', padx=20, pady=20)

# --- Gerenciamento de Departamentos ---
deptos_main_frame = ctk.CTkFrame(agentes_frame, fg_color="transparent")
deptos_main_frame.pack(pady=10, padx=10, fill="both", expand=True)

ctk.CTkLabel(deptos_main_frame, text="Gerenciar Departamentos (via API)", font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(5, 10))

add_depto_frame = ctk.CTkFrame(deptos_main_frame, fg_color="transparent")
add_depto_frame.pack(pady=5, fill="x")
ctk.CTkLabel(add_depto_frame, text="Nome do Novo Departamento:").pack(side="left", padx=5)
entry_novo_depto = ctk.CTkEntry(add_depto_frame, width=250, fg_color="#2E5A88")
entry_novo_depto.pack(side="left", padx=5)
botao_adicionar_depto = ctk.CTkButton(add_depto_frame, text="Criar Departamento", command=adicionar_departamento, font=ctk.CTkFont(weight="bold"))
botao_adicionar_depto.pack(side="left", padx=10)

list_depto_frame = ctk.CTkFrame(deptos_main_frame, fg_color="transparent")
list_depto_frame.pack(pady=10, fill="both", expand=True)
ctk.CTkLabel(list_depto_frame, text="Departamentos Existentes:", anchor="w").pack(fill="x", padx=5)
listbox_deptos = tk.Listbox(list_depto_frame, height=5, bg="#2a2d2e", fg="#FFFFFF", font=("Arial", 12), selectbackground="#4A90E2", selectforeground="#FFFFFF")
listbox_deptos.pack(fill="both", expand=True, padx=5, pady=5)
listbox_depto_scroll = ctk.CTkScrollbar(list_depto_frame, orientation="vertical", command=listbox_deptos.yview)
listbox_depto_scroll.pack(side="right", fill="y")
listbox_deptos.configure(yscrollcommand=listbox_depto_scroll.set)

# --- INICIALIZAÇÃO ---
if __name__ == "__main__":
    init_db()
    _atualizar_checkboxes_departamentos()
    popular_lista_departamentos()
    atualizar_ultimos_lancamentos()
    atualizar_dashboard()
    app.mainloop()
